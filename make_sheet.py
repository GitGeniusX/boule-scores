"""Build a ready-to-use Google Sheets template (.xlsx) for Boule Score Tracker.

The user downloads the file, uploads it to Google Drive, opens with Google Sheets,
then follows the instructions tab to wire up the Apps Script webhook.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---- Palette (matches app's warm terrain theme, but readable in Sheets) ----
TERRA = "C14A1F"   # primary accent
CREAM = "FBF5E3"   # light surface
SAND = "F2E6C9"    # page background feel
INK = "2A1F10"     # dark text
MUTED = "6E5A3B"   # secondary text
GREEN = "1F6F5C"   # team A
BRONZE = "B05A00"  # team B
HEADER_BG = "2A1F10"
HEADER_TEXT = "FBF5E3"

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=HEADER_TEXT)
FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=INK)
FONT_SUB = Font(name="Calibri", size=12, color=MUTED, italic=True)
FONT_BODY = Font(name="Calibri", size=11, color=INK)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True, color=INK)
FONT_STEP = Font(name="Calibri", size=14, bold=True, color=TERRA)
FONT_CODE = Font(name="Consolas", size=10, color=INK)
FONT_H2 = Font(name="Calibri", size=14, bold=True, color=INK)

thin = Side(border_style="thin", color="C9B287")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- Sheet 1: Instruktioner ----------------------------------------------
def build_instructions(ws):
    ws.title = "Instruktioner"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 80])

    # title row
    ws.row_dimensions[1].height = 10
    ws["B2"] = "Boule Score Tracker"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 32

    ws["B3"] = "Setup-guide för din egen Google Sheet + Apps Script-webhook"
    ws["B3"].font = FONT_SUB
    ws.row_dimensions[3].height = 22

    ws.row_dimensions[4].height = 10

    sections = [
        ("Så här använder du den här mallen", None),
        (None, (
            "1.  Ladda upp den här .xlsx-filen till din Google Drive.\n"
            "2.  Högerklicka på filen → \"Öppna med\" → \"Google Kalkylark\".\n"
            "3.  I Google Kalkylark: \"Arkiv\" → \"Spara som Google Sheets\" — du får en riktig Google-mall som är kopplad till ditt konto.\n"
            "4.  Radera .xlsx-filen om du vill, den behövs inte längre.\n"
            "5.  Följ stegen nedan för att koppla den till Boule-appen."
        )),
        (None, None),
        ("Steg 1 — Öppna Apps Script", None),
        (None, (
            "I ditt nya Google Sheet: Menyn \"Tillägg\" → \"Apps Script\" "
            "(på engelska: Extensions → Apps Script).\n\n"
            "Ett nytt Apps Script-projekt öppnas med en fil som heter Code.gs.\n"
            "Radera allt innehåll i Code.gs."
        )),
        (None, None),
        ("Steg 2 — Klistra in koden", None),
        (None, (
            "Öppna fliken \"Apps Script-kod\" i detta kalkylblad och kopiera ALL kod därifrån.\n"
            "Klistra in i Code.gs. Klicka spara-ikonen (eller Ctrl/Cmd+S) och döp projektet "
            "till \"Boule webhook\"."
        )),
        (None, None),
        ("Steg 3 — Deploy som webbapp", None),
        (None, (
            "1.  I Apps Script: klicka \"Deploy\" (uppe till höger) → \"New deployment\".\n"
            "2.  Klicka kugghjulet intill \"Select type\" → välj \"Web app\".\n"
            "3.  Fyll i:\n"
            "       • Description:  Boule webhook v1\n"
            "       • Execute as:   Me  (din Google-adress)\n"
            "       • Who has access:  Anyone   ← VIKTIGT, annars går det inte att posta\n"
            "4.  Klicka \"Deploy\".\n"
            "5.  Första gången frågar Google om behörigheter:\n"
            "       • \"Authorize access\" → välj ditt konto\n"
            "       • \"Advanced\" → \"Go to Boule webhook (unsafe)\" → \"Allow\"\n"
            "     (det är ditt eget skript — \"unsafe\" betyder bara att Google inte granskat det)\n"
            "6.  Kopiera \"Web app URL\" — den ser ut som:\n"
            "       https://script.google.com/macros/s/AKfyc.../exec"
        )),
        (None, None),
        ("Steg 4 — Koppla Boule-appen", None),
        (None, (
            "1.  Öppna appen:  https://gitgeniusx.github.io/boule-scores/\n"
            "2.  Klicka kugghjulet (inställningar) uppe till höger.\n"
            "3.  Metod:  \"Egen webhook-URL (t.ex. Google Apps Script)\"\n"
            "4.  Webhook-URL:  klistra in URL:en från steg 3.\n"
            "5.  Sätt gärna en standardplats (t.ex. \"Sexdrega gräsmatta\").\n"
            "6.  Bocka i \"Använd telefonens plats (GPS) vid sparning\" om du vill ha koordinater.\n"
            "7.  Klicka Spara."
        )),
        (None, None),
        ("Steg 5 — Testa", None),
        (None, (
            "Spela en snabb match i appen (höj målet till 11 via \"Mål\"-knappen för snabbtest).\n"
            "När match är slut: klicka \"Spara match\".\n"
            "Gå till fliken \"Matcher\" i detta Sheet — en rad ska dyka upp med all data.\n"
            "Toasten \"skickat till Google Sheet\" i appen är normalt."
        )),
        (None, None),
        ("Uppdatera skriptet senare (behåll samma URL)", None),
        (None, (
            "Apps Script ger en NY URL varje gång du gör \"New deployment\". För att behålla "
            "samma URL:\n\n"
            "    Deploy → Manage deployments → pennan på befintlig deploy → "
            "Version: New version → Deploy."
        )),
        (None, None),
        ("Felsökning", None),
        (None, (
            "• \"HTTP 401\" eller matchen sparas inte:  Deploy-inställningen \"Who has access\" är "
            "inte satt till \"Anyone\". Gör om Steg 3.\n"
            "• Raden dyker inte upp:  kolla att du är i rätt Sheet — skriptet skriver till det Sheet "
            "det skapades i. Titta efter ett ark som heter \"Matcher\".\n"
            "• CORS-fel i konsolen:  normalt. Appen skickar med no-cors när URL är Apps Script. "
            "Raden skrivs ändå.\n"
            "• Ingen GPS:  kräver HTTPS + tillstånd i telefonen. Fungerar på appens live-URL.\n"
            "• Platsen är tom i Sheet:  GPS-timeout eller tillstånd nekat. Använd standardplats-"
            "fältet i appens inställningar som fallback."
        )),
    ]

    row = 5
    for header, body in sections:
        if header:
            ws.cell(row=row, column=2, value=header).font = FONT_STEP
            ws.row_dimensions[row].height = 24
            row += 1
        if body:
            cell = ws.cell(row=row, column=2, value=body)
            cell.font = FONT_BODY
            cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            # approximate height: 15pt per line
            lines = body.count("\n") + 1
            ws.row_dimensions[row].height = max(18, lines * 15 + 4)
            row += 1
        row += 1  # spacer


# ---- Sheet 2: Apps Script-kod --------------------------------------------
APPS_SCRIPT_CODE = '''/**
 * Boule Score Tracker — Google Sheets webhook
 * Tar emot POST-JSON från Boule-appen.
 * Skriver till tre flikar: Matcher, Lagsättningar, Spelare.
 */

const MATCH_SHEET = 'Matcher';
const LINEUP_SHEET = 'Lagsättningar';
const PLAYERS_SHEET = 'Spelare';

const MATCH_HEADERS = [
  'Match-ID',
  'Sparad',
  'Startade',
  'Slutade',
  'Tidszon',
  'Plats',
  'Latitud',
  'Longitud',
  'GPS-noggrannhet (m)',
  'Mål',
  'Lag 1',
  'Spelare Lag 1',
  'Poäng 1',
  'Lag 2',
  'Spelare Lag 2',
  'Poäng 2',
  'Vinnare',
  'Antal omgångar',
  'Omgångar (JSON)',
];

const LINEUP_HEADERS = ['Match-ID', 'Sparad', 'Lag', 'Lagnamn', 'Spelare'];
const PLAYER_HEADERS = ['Spelare', 'Matcher', 'Vinster', 'Senast sedd'];

function doPost(e) {
  try {
    const raw = e && e.postData && e.postData.contents;
    if (!raw) throw new Error('Tom request body');
    const data = JSON.parse(raw);

    const matchId = (data.endedAt || data.savedAt || new Date().toISOString()) + '-' +
      Math.random().toString(36).slice(2, 7);

    writeMatch_(matchId, data);
    writeLineups_(matchId, data);
    upsertPlayers_(data);

    return json_({ ok: true, matchId: matchId });
  } catch (err) {
    return json_({ ok: false, error: String(err) });
  }
}

function doGet() {
  return json_({
    ok: true,
    service: 'Boule webhook v2',
    usage: 'POST JSON från Boule-appen till denna URL',
  });
}

function writeMatch_(matchId, data) {
  const sheet = getOrCreateSheet_(MATCH_SHEET, MATCH_HEADERS);
  const t1 = (data.teams && data.teams[0]) || {};
  const t2 = (data.teams && data.teams[1]) || {};
  sheet.appendRow([
    matchId,
    data.savedAt || new Date().toISOString(),
    data.startedAt || '',
    data.endedAt || '',
    data.timezone || '',
    data.place || '',
    data.location ? data.location.lat : '',
    data.location ? data.location.lng : '',
    data.location ? data.location.accuracy : '',
    data.target || '',
    t1.name || '',
    (t1.players || []).join(', '),
    t1.score || 0,
    t2.name || '',
    (t2.players || []).join(', '),
    t2.score || 0,
    data.winner || '',
    (data.rounds && data.rounds.length) || 0,
    JSON.stringify(data.rounds || []),
  ]);
}

function writeLineups_(matchId, data) {
  const sheet = getOrCreateSheet_(LINEUP_SHEET, LINEUP_HEADERS);
  const saved = data.savedAt || new Date().toISOString();
  (data.teams || []).forEach((team, idx) => {
    (team.players || []).forEach((p) => {
      sheet.appendRow([matchId, saved, 'Lag ' + (idx + 1), team.name || '', p]);
    });
  });
}

function upsertPlayers_(data) {
  const sheet = getOrCreateSheet_(PLAYERS_SHEET, PLAYER_HEADERS);
  const values = sheet.getDataRange().getValues(); // inkl header
  const indexByName = {};
  for (let r = 1; r < values.length; r++) {
    indexByName[String(values[r][0]).toLowerCase()] = r + 1; // 1-indexerat radnummer
  }
  const saved = data.savedAt || new Date().toISOString();
  const allPlayers = [];
  (data.teams || []).forEach((team) => {
    (team.players || []).forEach((p) => {
      allPlayers.push({ name: p, team: team.name });
    });
  });
  const winner = data.winner;

  allPlayers.forEach((pl) => {
    const key = pl.name.toLowerCase();
    const existingRow = indexByName[key];
    if (existingRow) {
      const matches = Number(sheet.getRange(existingRow, 2).getValue()) || 0;
      const wins = Number(sheet.getRange(existingRow, 3).getValue()) || 0;
      sheet.getRange(existingRow, 2).setValue(matches + 1);
      if (pl.team === winner) sheet.getRange(existingRow, 3).setValue(wins + 1);
      sheet.getRange(existingRow, 4).setValue(saved);
    } else {
      sheet.appendRow([pl.name, 1, pl.team === winner ? 1 : 0, saved]);
      indexByName[key] = sheet.getLastRow();
    }
  });
}

function getOrCreateSheet_(name, headers) {
  const ss = SpreadsheetApp.getActiveSpreadsheet();
  let sheet = ss.getSheetByName(name);
  if (!sheet) sheet = ss.insertSheet(name);
  if (sheet.getLastRow() === 0) {
    sheet.appendRow(headers);
    sheet.getRange(1, 1, 1, headers.length).setFontWeight('bold');
    sheet.setFrozenRows(1);
  }
  return sheet;
}

function json_(obj) {
  return ContentService.createTextOutput(JSON.stringify(obj)).setMimeType(
    ContentService.MimeType.JSON,
  );
}
'''


def build_code(ws):
    ws.title = "Apps Script-kod"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 110])

    ws["B2"] = "Apps Script-kod"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 32

    ws["B3"] = (
        "Markera cellen nedan, kopiera (Ctrl/Cmd+C) och klistra in i Code.gs i Apps Script. "
        "Radera det befintliga innehållet i Code.gs först."
    )
    ws["B3"].font = FONT_SUB
    ws["B3"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 32

    ws["B5"] = APPS_SCRIPT_CODE
    ws["B5"].font = FONT_CODE
    ws["B5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B5"].fill = PatternFill("solid", fgColor=CREAM)
    ws["B5"].border = BORDER
    # ~17pt per line of monospace
    ws.row_dimensions[5].height = APPS_SCRIPT_CODE.count("\n") * 14 + 20


# ---- Sheet 3: Matcher (dit webhooken skriver) ----------------------------
MATCH_HEADERS = [
    "Match-ID",
    "Sparad",
    "Startade",
    "Slutade",
    "Tidszon",
    "Plats",
    "Latitud",
    "Longitud",
    "GPS-noggrannhet (m)",
    "Mål",
    "Lag 1",
    "Spelare Lag 1",
    "Poäng 1",
    "Lag 2",
    "Spelare Lag 2",
    "Poäng 2",
    "Vinnare",
    "Antal omgångar",
    "Omgångar (JSON)",
]

EXAMPLE_ROW = [
    "2026-04-18T20:15:30.000Z-ab12c",
    "2026-04-18T20:15:30.000Z",
    "2026-04-18T19:45:12.000Z",
    "2026-04-18T20:15:28.000Z",
    "Europe/Stockholm",
    "Sexdrega, Svenljunga, Sverige",
    57.82015,
    13.19043,
    15,
    13,
    "Tjejerna",
    "Anna, Mia",
    13,
    "Killarna",
    "Erik, Johan",
    9,
    "Tjejerna",
    11,
    '[{"n":1,"team":"Tjejerna","points":2,"tally":[2,0]}, …]',
]


def build_matches(ws):
    ws.title = "Matcher"
    ws.sheet_view.showGridLines = False

    # Header row (row 1, som Apps Script förväntar sig)
    for col, header in enumerate(MATCH_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = FONT_HEADER
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Example row — gray italic
    for col, value in enumerate(EXAMPLE_ROW, start=1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = Font(name="Calibri", size=10, italic=True, color="999999")
        cell.border = BORDER

    # Frozen header
    ws.freeze_panes = "A2"

    # Number formats on example row (and future rows via column default)
    # Kolumner: A=ID, B=Sparad, C=Startade, D=Slutade, E=Tidszon, F=Plats,
    # G=Lat, H=Lng, I=Acc, J=Mål, K=Lag1, L=Spelare1, M=Po1, N=Lag2, O=Spelare2,
    # P=Po2, Q=Vinnare, R=Antal, S=JSON
    ws.cell(row=2, column=7).number_format = "0.00000"   # Lat
    ws.cell(row=2, column=8).number_format = "0.00000"   # Lng
    ws.cell(row=2, column=9).number_format = "0"         # Acc
    for c in (10, 13, 16, 18):
        ws.cell(row=2, column=c).number_format = "0"

    # Column widths (19 kolumner)
    widths = [28, 22, 22, 22, 18, 24, 11, 11, 10, 8, 16, 22, 10, 16, 22, 10, 16, 10, 50]
    set_col_widths(ws, widths)
    ws.row_dimensions[1].height = 30

    # Note row below the example
    note = ws.cell(
        row=4,
        column=1,
        value=(
            "Raden ovan är bara ett exempel — du kan radera den. "
            "När Apps Script-webhooken är igång kommer varje avslutad match i appen skapa en ny rad här."
        ),
    )
    note.font = FONT_SUB
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(MATCH_HEADERS))
    note.alignment = Alignment(wrap_text=True, vertical="top")


# ---- Sheet 4: Sammanfattning (formler) -----------------------------------
def build_lineups(ws):
    ws.title = "Lagsättningar"
    ws.sheet_view.showGridLines = False

    headers = ["Match-ID", "Sparad", "Lag", "Lagnamn", "Spelare"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    examples = [
        ["2026-04-18T20:15:30.000Z-ab12c", "2026-04-18T20:15:30.000Z", "Lag 1", "Tjejerna", "Anna"],
        ["2026-04-18T20:15:30.000Z-ab12c", "2026-04-18T20:15:30.000Z", "Lag 1", "Tjejerna", "Mia"],
        ["2026-04-18T20:15:30.000Z-ab12c", "2026-04-18T20:15:30.000Z", "Lag 2", "Killarna", "Erik"],
        ["2026-04-18T20:15:30.000Z-ab12c", "2026-04-18T20:15:30.000Z", "Lag 2", "Killarna", "Johan"],
    ]
    for i, row in enumerate(examples, start=2):
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(name="Calibri", size=10, italic=True, color="999999")
            c.border = BORDER

    ws.freeze_panes = "A2"
    set_col_widths(ws, [30, 22, 10, 16, 18])
    ws.row_dimensions[1].height = 30

    note = ws.cell(
        row=7,
        column=1,
        value=(
            "En rad per spelare och match. Fylls på automatiskt av webhooken. "
            "Du kan ta bort exempelraderna ovan."
        ),
    )
    note.font = FONT_SUB
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=5)


def build_players(ws):
    ws.title = "Spelare"
    ws.sheet_view.showGridLines = False

    headers = ["Spelare", "Matcher", "Vinster", "Senast sedd"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    examples = [
        ["Anna", 1, 1, "2026-04-18T20:15:30.000Z"],
        ["Erik", 1, 0, "2026-04-18T20:15:30.000Z"],
        ["Johan", 1, 0, "2026-04-18T20:15:30.000Z"],
        ["Mia", 1, 1, "2026-04-18T20:15:30.000Z"],
    ]
    for i, row in enumerate(examples, start=2):
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(name="Calibri", size=10, italic=True, color="999999")
            c.border = BORDER

    ws.freeze_panes = "A2"
    set_col_widths(ws, [20, 10, 10, 24])
    ws.row_dimensions[1].height = 30

    note = ws.cell(
        row=7,
        column=1,
        value=(
            "Webhooken underhåller denna flik automatiskt. Varje spelare får en rad "
            "med totalt antal matcher, vinster och när de senast spelade."
        ),
    )
    note.font = FONT_SUB
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=4)


def build_summary(ws):
    ws.title = "Sammanfattning"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 28, 14, 14, 14, 14])

    ws.row_dimensions[1].height = 10
    ws["B2"] = "Sammanfattning"
    ws["B2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 32
    ws["B3"] = "Automatiskt uppdaterad när nya matcher sparas i \"Matcher\"-fliken."
    ws["B3"].font = FONT_SUB
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 10

    # Kolumnkarta för Matcher-fliken (19 kolumner):
    # A=ID, B=Sparad, C=Startade, D=Slutade, E=Tidszon, F=Plats,
    # G=Lat, H=Lng, I=Acc, J=Mål, K=Lag1, L=Spelare1, M=Po1,
    # N=Lag2, O=Spelare2, P=Po2, Q=Vinnare, R=Antal omg, S=JSON

    # KPIs
    kpis = [
        ("Totalt spelade matcher", '=COUNTA(Matcher!Q:Q)-1'),
        ("Tjejernas segrar", '=COUNTIF(Matcher!Q:Q,"Tjejerna")'),
        ("Killarnas segrar", '=COUNTIF(Matcher!Q:Q,"Killarna")'),
        ("Snitt omgångar per match", '=IFERROR(AVERAGE(Matcher!R2:R1000),0)'),
        ("Antal unika spelare", '=IFERROR(COUNTA(Spelare!A:A)-1,0)'),
    ]
    for i, (label, formula) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=2, value=label).font = FONT_BODY
        v = ws.cell(row=r, column=3, value=formula)
        v.font = FONT_BODY_BOLD
        v.alignment = Alignment(horizontal="right")
        if "AVERAGE" in formula:
            v.number_format = "0.0"
        else:
            v.number_format = "#,##0"
        ws.row_dimensions[r].height = 20

    # Spacing
    r = 10
    ws.cell(row=r, column=2, value="Per lag").font = FONT_H2
    ws.row_dimensions[r].height = 24

    # Header row for per-team summary
    r += 1
    headers = ["Lag", "Segrar", "Totala poäng gjorda", "Senast spelad"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=r, column=2 + i, value=h)
        cell.font = FONT_HEADER
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[r].height = 22

    team_rows = [
        (
            "Tjejerna",
            '=COUNTIF(Matcher!Q:Q,B{r})',
            # Summa: om lag 1 heter så → summera poäng 1, om lag 2 heter så → summera poäng 2
            '=SUMIF(Matcher!K:K,B{r},Matcher!M:M)+SUMIF(Matcher!N:N,B{r},Matcher!P:P)',
            '=IFERROR(MAXIFS(Matcher!D:D,Matcher!Q:Q,B{r}),"—")',
        ),
        (
            "Killarna",
            '=COUNTIF(Matcher!Q:Q,B{r})',
            '=SUMIF(Matcher!K:K,B{r},Matcher!M:M)+SUMIF(Matcher!N:N,B{r},Matcher!P:P)',
            '=IFERROR(MAXIFS(Matcher!D:D,Matcher!Q:Q,B{r}),"—")',
        ),
    ]
    for i, (team, wins_f, pts_f, last_f) in enumerate(team_rows):
        r2 = r + 1 + i
        ws.cell(row=r2, column=2, value=team).font = FONT_BODY_BOLD
        w = ws.cell(row=r2, column=3, value=wins_f.format(r=r2))
        w.number_format = "0"
        w.alignment = Alignment(horizontal="right")
        p = ws.cell(row=r2, column=4, value=pts_f.format(r=r2))
        p.number_format = "0"
        p.alignment = Alignment(horizontal="right")
        d = ws.cell(row=r2, column=5, value=last_f.format(r=r2))
        d.alignment = Alignment(horizontal="right")
        for c in range(2, 6):
            ws.cell(row=r2, column=c).border = BORDER
        ws.row_dimensions[r2].height = 20

    # Senaste 5 matcher
    r = r + 4
    ws.cell(row=r, column=2, value="Senaste 5 matcher").font = FONT_H2
    ws.row_dimensions[r].height = 24
    r += 1

    sub_headers = ["Datum", "Lag 1", "Poäng", "Lag 2", "Vinnare"]
    for i, h in enumerate(sub_headers):
        cell = ws.cell(row=r, column=2 + i, value=h)
        cell.font = FONT_HEADER
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[r].height = 22

    # Formula: använd IFERROR + INDEX för att plocka senaste matcherna (omvänd ordning)
    # Antag att Matcher!B2:B1000 är "Sparad" (kolumn B nu när A är ID)
    for i in range(5):
        r2 = r + 1 + i
        # offset: det finns n matcher → plocka rad (n - i + 1) i Matcher (header är rad 1)
        n_matches = 'COUNTA(Matcher!B:B)-1'
        offset = f'{n_matches}-{i}+1'  # 1-indexerat, header är rad 1
        # Datum=D (Slutade), Lag1=K, Po1=M, Po2=P, Lag2=N, Vinnare=Q
        ws.cell(row=r2, column=2, value=f'=IFERROR(INDEX(Matcher!D:D,{offset}),"—")')
        ws.cell(row=r2, column=3, value=f'=IFERROR(INDEX(Matcher!K:K,{offset}),"—")')
        ws.cell(row=r2, column=4, value=f'=IFERROR(INDEX(Matcher!M:M,{offset})&" – "&INDEX(Matcher!P:P,{offset}),"—")')
        ws.cell(row=r2, column=5, value=f'=IFERROR(INDEX(Matcher!N:N,{offset}),"—")')
        ws.cell(row=r2, column=6, value=f'=IFERROR(INDEX(Matcher!Q:Q,{offset}),"—")')
        for c in range(2, 7):
            cell = ws.cell(row=r2, column=c)
            cell.border = BORDER
            cell.font = FONT_BODY
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r2].height = 20

    # Footer note
    ws.cell(row=r2 + 3, column=2, value=(
        "Tips: ändra lagnamnen i cell B12 och B13 om ni döper om lagen i appen — "
        "formlerna uppdateras automatiskt."
    )).font = FONT_SUB
    ws.merge_cells(start_row=r2 + 3, start_column=2, end_row=r2 + 3, end_column=6)


# ---- Build workbook ------------------------------------------------------
def main():
    wb = Workbook()

    # Sheet 1 (default)
    build_instructions(wb.active)

    build_code(wb.create_sheet())
    build_matches(wb.create_sheet())
    build_lineups(wb.create_sheet())
    build_players(wb.create_sheet())
    build_summary(wb.create_sheet())

    # Sheet order: Instruktioner, Apps Script-kod, Matcher, Lagsättningar, Spelare, Sammanfattning

    out = "/home/user/workspace/boule-scores/Boule-matcher-mall.xlsx"
    wb.save(out)
    print(f"saved -> {out}")


if __name__ == "__main__":
    main()
